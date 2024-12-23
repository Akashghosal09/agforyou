#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Dec 13 17:57:55 2024

@author: akashghoshal
"""


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, date
import openpyxl

def read_actor_names(filename):
    """Reads actor names from an Excel file."""
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    actor_names = [cell.value for cell in sheet['A']]  # Assuming names are in column A
    return actor_names

def get_actor_birthdate(driver, actor_name):
    """Searches Wikipedia for the actor's birthdate using Selenium."""
    url = f"https://en.wikipedia.org/wiki/{actor_name}"
    driver.get(url)

    try:
        # Wait for the birthdate element to be present (adjust XPATH as needed)
        birthdate_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//span[@class='bday']"))
        )
        birthdate = birthdate_element.text
    except:
        birthdate = "Not Found"

    return birthdate

def calculate_age(birthdate):
    """Calculates the age based on the birthdate."""
    if birthdate == "Not Found":
        return "N/A"
    try:
        birthdate_obj = datetime.strptime(birthdate, "%Y-%m-%d") 
        today = date.today()
        age = today.year - birthdate_obj.year 
        return age
    except ValueError:
        print(f"Invalid birthdate format: {birthdate}")
        return "Invalid"

def save_data_to_excel(actor_names, birthdates, ages, filename):
    """Saves actor names, birthdates, and ages to a new Excel file."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = "Actor Name"
    sheet['B1'] = "Date of Birth"
    sheet['C1'] = "Age (Years)"  # Column heading for clarity
    for i, actor_name in enumerate(actor_names):
        sheet.cell(row=i+2, column=1, value=actor_name)
        sheet.cell(row=i+2, column=2, value=birthdates[i])
        sheet.cell(row=i+2, column=3, value=ages[i])
    workbook.save(filename)

if __name__ == "__main__":
    # Set up Selenium driver
    driver = webdriver.Chrome()  # Replace with your preferred browser (e.g., Firefox, Edge)

    actor_names = read_actor_names("actor_names.xlsx")
    birthdates = []
    ages = []

    for name in actor_names:
        birthdate = get_actor_birthdate(driver, name)
        birthdates.append(birthdate)
        age = calculate_age(birthdate)
        ages.append(age)

    save_data_to_excel(actor_names, birthdates, ages, "actor_birthdates.xlsx")

    # Close the browser
    driver.quit()