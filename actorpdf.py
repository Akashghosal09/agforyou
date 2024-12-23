#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Dec 13 20:47:47 2024

@author: akashghoshal
"""

import openpyxl
import requests
from bs4 import BeautifulSoup
from datetime import date, datetime

def read_actor_names(filename):
    """Reads actor names from an Excel file."""
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    actor_names = [cell.value for cell in sheet['A']]  # Assuming names are in column A
    return actor_names

def get_actor_birthdate(actor_name):
    """Searches Wikipedia for the actor's birthdate."""
    url = f"https://en.wikipedia.org/wiki/{actor_name}"
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for bad status codes
        soup = BeautifulSoup(response.text, "html.parser")
        # Adapt this part based on Wikipedia's HTML structure
        birthdate_tag = soup.find("span", class_="bday") 
        if birthdate_tag:
            return birthdate_tag.text
        else:
            return "Not Found"
    except (requests.exceptions.RequestException, AttributeError) as e:
        print(f"Error processing {actor_name}: {e}")
        return "Error"

def calculate_age(birthdate):
    """Calculates the age based on the birthdate."""
    if birthdate == "Not Found" or birthdate == "Error":
        return "N/A"
    try:
        birthdate_obj = datetime.strptime(birthdate, "%Y-%m-%d") 
        today = date.today()
        age = today.year - birthdate_obj.year - ((today.month, today.day) < (birthdate_obj.month, birthdate_obj.day))
        return age
    except ValueError:
        print(f"Invalid birthdate format: {birthdate}")
        return "Invalid"

def save_data_to_excel(actor_names, birthdates, ages, filename):
    """Saves actor names, birthdates, and ages to a new Excel file."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = "Actor Name"
    sheet['B1'] = "Date of Birth"
    sheet['C1'] = "Age"
    for i, actor_name in enumerate(actor_names):
        sheet.cell(row=i+2, column=1, value=actor_name)
        sheet.cell(row=i+2, column=2, value=birthdates[i])
        sheet.cell(row=i+2, column=3, value=ages[i])
    workbook.save(filename)

# **Add printing functionality (using a library like `pdfkit`)**
# import pdfkit 

if __name__ == "__main__":
    actor_names = read_actor_names("actor_names.xlsx")  # Replace with your input file
    birthdates = []
    ages = []
    for name in actor_names:
        birthdate = get_actor_birthdate(name)
        birthdates.append(birthdate)
        age = calculate_age(birthdate)
        ages.append(age)

# **Add printing functionality (using a library like `pdfkit`)**
url = f"https://en.wikipedia.org/wiki/{name}"
pdfkit.from_url(url, f"{name}.pdf") 

save_data_to_excel(actor_names, birthdates, ages, "actor_birthdates.xlsx")
       
       # Close the browser
driver.quit() 
       