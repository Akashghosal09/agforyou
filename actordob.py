#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Dec 13 19:52:33 2024

@author: akashghoshal
"""

import openpyxl
import requests
from bs4 import BeautifulSoup

def read_actor_names(filename):
    """Reads actor names from an Excel file."""
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    actor_names = [cell.value for cell in sheet['A']]  # Assuming names are in column A
    return actor_names

def get_actor_birthdate(actor_name):
    """Searches Wikipedia for the actor's birthdate."""
    url = f"https://en.wikipedia.org/wiki/{actor_name}"
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for bad status codes
        soup = BeautifulSoup(response.text, "html.parser")
        # Adapt this part based on Wikipedia's HTML structure
        birthdate_tag = soup.find("span", class_="bday") 
        if birthdate_tag:
            return birthdate_tag.text
        else:
            return "Not Found"
    except (requests.exceptions.RequestException, AttributeError) as e:
        print(f"Error processing {actor_name}: {e}")
        return "Error"

def save_data_to_excel(actor_names, birthdates, filename):
    """Saves actor names and birthdates to a new Excel file."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = "Actor Name"
    sheet['B1'] = "Date of Birth"
    for i, actor_name in enumerate(actor_names):
        sheet.cell(row=i+2, column=1, value=actor_name)
        sheet.cell(row=i+2, column=2, value=birthdates[i])
    workbook.save(filename)

if __name__ == "__main__":
    actor_names = read_actor_names("actor_names.xlsx")  # Replace with your input file
    birthdates = []
    for name in actor_names:
        birthdate = get_actor_birthdate(name)
        birthdates.append(birthdate)
    save_data_to_excel(actor_names, birthdates, "actor_birthdates.xlsx")